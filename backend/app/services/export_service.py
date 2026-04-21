from __future__ import annotations

import csv
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO, StringIO

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.patches import Rectangle
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Dog, Worklog
from app.services.analytics_service import (
    get_analytics_summary,
    get_weekly_analytics,
    get_weekly_compare,
)

plt.rcParams["font.family"] = "DejaVu Sans"


THIN_GRAY = Side(style="thin", color="D1D5DB")
THIN_DARK = Side(style="thin", color="111827")
BORDER_ALL = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
BORDER_DARK = Border(left=THIN_DARK, right=THIN_DARK, top=THIN_DARK, bottom=THIN_DARK)

FILL_TITLE = PatternFill("solid", fgColor="0F172A")
FILL_SECTION = PatternFill("solid", fgColor="E5E7EB")
FILL_HEADER = PatternFill("solid", fgColor="CBD5E1")
FILL_SUBHEADER = PatternFill("solid", fgColor="E2E8F0")
FILL_INFO = PatternFill("solid", fgColor="EFF6FF")
FILL_GOOD = PatternFill("solid", fgColor="ECFDF5")
FILL_WARN = PatternFill("solid", fgColor="FFF7ED")
FILL_WEEK_A = PatternFill("solid", fgColor="FDF2F8")
FILL_WEEK_B = PatternFill("solid", fgColor="FFF7ED")
FILL_TOTAL = PatternFill("solid", fgColor="ECFDF5")
FILL_NAME = PatternFill("solid", fgColor="EFF6FF")
FILL_SUMMARY = PatternFill("solid", fgColor="F8FAFC")

FONT_TITLE = Font(color="FFFFFF", bold=True, size=14)
FONT_SECTION = Font(color="111827", bold=True, size=12)
FONT_HEADER = Font(color="111827", bold=True)
FONT_BOLD = Font(color="111827", bold=True)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _to_float(value: Decimal | float | int | None) -> float:
    if value is None:
        return 0.0
    return float(value)


def _start_of_week(value: date) -> date:
    return value - timedelta(days=value.weekday())


def _is_operational_dog(dog: Dog) -> bool:
    lifecycle = str(
        dog.lifecycle_status.value if hasattr(dog.lifecycle_status, "value") else dog.lifecycle_status
    ).lower()
    return lifecycle not in {"archived", "deceased"}


def _autosize_columns(ws, min_width: int = 12, max_width: int = 36) -> None:
    widths: dict[int, int] = {}

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)) + 2)

    for column_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(column_idx)].width = max(min_width, min(width, max_width))


def _set_number_formats(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.00"


def _write_title(ws, title: str, subtitle: str | None = None, width_to_col: int = 9) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width_to_col)
    cell = ws.cell(row=1, column=1, value=title)
    cell.fill = FILL_TITLE
    cell.font = FONT_TITLE
    cell.alignment = ALIGN_LEFT
    cell.border = BORDER_DARK

    for col in range(1, width_to_col + 1):
        ws.cell(row=1, column=col).border = BORDER_DARK
        ws.cell(row=1, column=col).fill = FILL_TITLE

    current_row = 2
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width_to_col)
        sub = ws.cell(row=2, column=1, value=subtitle)
        sub.alignment = ALIGN_LEFT
        sub.border = BORDER_ALL
        for col in range(1, width_to_col + 1):
            ws.cell(row=2, column=col).border = BORDER_ALL
        current_row = 3

    return current_row


def _merge_with_style(
    ws,
    start_row: int,
    start_col: int,
    end_row: int,
    end_col: int,
    value: str | None,
    fill: PatternFill,
    font: Font,
    alignment: Alignment,
    border: Border = BORDER_DARK,
) -> None:
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    anchor = ws.cell(row=start_row, column=start_col, value=value)
    anchor.fill = fill
    anchor.font = font
    anchor.alignment = alignment
    anchor.border = border

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = border
            if r == start_row and c == start_col:
                cell.font = font
                cell.alignment = alignment


def _fetch_operational_dogs(db: Session) -> list[Dog]:
    dogs = list(
        db.execute(
            select(Dog).order_by(
                Dog.kennel_row.asc(),
                Dog.kennel_block.asc(),
                Dog.home_slot.asc(),
                Dog.name.asc(),
            )
        ).scalars().all()
    )
    return [dog for dog in dogs if _is_operational_dog(dog)]


def _fetch_worklogs_in_range(db: Session, date_from: date, date_to: date) -> list[Worklog]:
    return list(
        db.execute(
            select(Worklog)
            .where(Worklog.work_date >= date_from, Worklog.work_date <= date_to)
            .order_by(Worklog.work_date.asc(), Worklog.dog_id.asc(), Worklog.id.asc())
        ).scalars().all()
    )


def _comparison_hint(metric: str, delta: float | int) -> str:
    if delta == 0:
        return "No change"

    if metric in {"total_km", "worked_dogs", "avg_km_per_worked_dog"}:
        return "Improved" if delta > 0 else "Lower"

    if metric in {"high_risk_dogs", "moderate_risk_dogs", "underused_dogs"}:
        return "Reduced" if delta < 0 else "Increased"

    return "Changed"


def _safe_ratio(numerator: float, denominator: float) -> float | None:
    if denominator <= 0:
        return None
    return round(numerator / denominator, 2)


def _week_fill(index: int) -> PatternFill:
    return FILL_WEEK_A if index % 2 == 0 else FILL_WEEK_B


def _weekly_compare_rows(comparison):
    return [
        ("Total km", comparison.week_a.total_km, comparison.week_b.total_km, comparison.delta.total_km, "total_km"),
        ("Worked dogs", comparison.week_a.worked_dogs, comparison.week_b.worked_dogs, comparison.delta.worked_dogs, "worked_dogs"),
        ("Avg km / worked dog", comparison.week_a.avg_km_per_worked_dog, comparison.week_b.avg_km_per_worked_dog, comparison.delta.avg_km_per_worked_dog, "avg_km_per_worked_dog"),
        ("High risk", comparison.week_a.high_risk_dogs, comparison.week_b.high_risk_dogs, comparison.delta.high_risk_dogs, "high_risk_dogs"),
        ("Moderate risk", comparison.week_a.moderate_risk_dogs, comparison.week_b.moderate_risk_dogs, comparison.delta.moderate_risk_dogs, "moderate_risk_dogs"),
        ("Underused", comparison.week_a.underused_dogs, comparison.week_b.underused_dogs, comparison.delta.underused_dogs, "underused_dogs"),
    ]


def _build_overview_sheet(wb: Workbook, date_from: date, date_to: date, weekly, summary, comparison) -> None:
    ws = wb.active
    ws.title = "Overview"

    row = _write_title(
        ws,
        title="Husky Tracking AI - Analytics Export",
        subtitle=f"Period: {date_from.isoformat()} -> {date_to.isoformat()} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        width_to_col=10,
    )

    cell = ws.cell(row=row, column=1, value="Summary")
    cell.fill = FILL_SECTION
    cell.font = FONT_SECTION
    cell.border = BORDER_ALL
    row += 1

    summary_rows = [
        ("Total km", summary.total_km),
        ("Worked dogs", summary.unique_worked_dogs),
        ("Avg km / worked dog", summary.avg_km_per_worked_dog),
        ("Worked dog-days", summary.total_worked_dog_days),
        ("Weeks in export", summary.weeks_count),
        (
            "Latest snapshot",
            f"H {summary.latest_week_snapshot.high_risk_dogs} / M {summary.latest_week_snapshot.moderate_risk_dogs} / U {summary.latest_week_snapshot.underused_dogs}"
            if summary.latest_week_snapshot
            else "N/A",
        ),
    ]

    for label, value in summary_rows:
        left = ws.cell(row=row, column=1, value=label)
        left.font = FONT_HEADER
        left.fill = FILL_SUBHEADER
        left.border = BORDER_ALL

        right = ws.cell(row=row, column=2, value=value)
        right.fill = FILL_INFO if label == "Latest snapshot" else FILL_SUMMARY
        right.border = BORDER_ALL
        right.alignment = ALIGN_LEFT
        row += 1

    row += 1
    section = ws.cell(row=row, column=1, value="Weekly summary")
    section.fill = FILL_SECTION
    section.font = FONT_SECTION
    section.border = BORDER_ALL
    row += 1

    headers = [
        "Week start",
        "Week end",
        "Week label",
        "Total km",
        "Worked dogs",
        "Avg km / worked dog",
        "High risk",
        "Moderate risk",
        "Underused",
    ]
    for col, header in enumerate(headers, start=1):
        h = ws.cell(row=row, column=col, value=header)
        h.fill = FILL_HEADER
        h.font = FONT_HEADER
        h.alignment = ALIGN_CENTER
        h.border = BORDER_ALL

    row += 1
    for index, item in enumerate(weekly.items):
        fill = _week_fill(index)
        values = [
            item.week_start.isoformat(),
            item.week_end.isoformat(),
            item.week_label,
            item.total_km,
            item.worked_dogs,
            item.avg_km_per_worked_dog,
            item.high_risk_dogs,
            item.moderate_risk_dogs,
            item.underused_dogs,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = BORDER_ALL
            cell.alignment = ALIGN_LEFT if col <= 3 else ALIGN_CENTER
            if col >= 4:
                cell.fill = fill
        row += 1

    if comparison is not None:
        row += 1
        section = ws.cell(row=row, column=1, value="Selected week comparison")
        section.fill = FILL_SECTION
        section.font = FONT_SECTION
        section.border = BORDER_ALL
        row += 1

        _merge_with_style(
            ws,
            row,
            1,
            row,
            5,
            "Week A = first selected week | Week B = second selected week | Change = Week B minus Week A",
            FILL_INFO,
            FONT_HEADER,
            ALIGN_LEFT,
            BORDER_ALL,
        )
        row += 1

        compare_headers = ["Metric", "Week A", "Week B", "Change", "Interpretation"]
        for col, header in enumerate(compare_headers, start=1):
            h = ws.cell(row=row, column=col, value=header)
            h.fill = FILL_HEADER
            h.font = FONT_HEADER
            h.alignment = ALIGN_CENTER
            h.border = BORDER_ALL

        row += 1
        for metric, week_a_value, week_b_value, delta_value, metric_key in _weekly_compare_rows(comparison):
            hint = _comparison_hint(metric_key, delta_value)
            values = [metric, week_a_value, week_b_value, delta_value, hint]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = BORDER_ALL
                cell.alignment = ALIGN_CENTER if col > 1 else ALIGN_LEFT
                if col == 4:
                    if metric_key in {"high_risk_dogs", "moderate_risk_dogs", "underused_dogs"}:
                        cell.fill = FILL_GOOD if delta_value < 0 else FILL_WARN if delta_value > 0 else FILL_SUMMARY
                    else:
                        cell.fill = FILL_GOOD if delta_value > 0 else FILL_WARN if delta_value < 0 else FILL_SUMMARY
            row += 1

    ws.freeze_panes = "A4"
    _set_number_formats(ws)
    _autosize_columns(ws, min_width=14, max_width=34)


def _build_weekly_comparison_sheet(wb: Workbook, comparison) -> None:
    ws = wb.create_sheet("Weekly Comparison")

    row = _write_title(
        ws,
        title="Weekly Comparison",
        subtitle=(f"{comparison.week_a.week_label} vs {comparison.week_b.week_label}" if comparison is not None else "No specific comparison selected"),
        width_to_col=6,
    )

    _merge_with_style(
        ws,
        row,
        1,
        row,
        5,
        "Week A = first selected week | Week B = second selected week | Change = Week B minus Week A",
        FILL_INFO,
        FONT_HEADER,
        ALIGN_LEFT,
        BORDER_ALL,
    )
    row += 1

    headers = ["Metric", "Week A", "Week B", "Change", "Interpretation"]
    for col, header in enumerate(headers, start=1):
        h = ws.cell(row=row, column=col, value=header)
        h.fill = FILL_HEADER
        h.font = FONT_HEADER
        h.alignment = ALIGN_CENTER
        h.border = BORDER_ALL
    row += 1

    if comparison is None:
        _merge_with_style(ws, row, 1, row, 5, "No weeks selected for comparison.", FILL_SUMMARY, FONT_HEADER, ALIGN_LEFT, BORDER_ALL)
    else:
        for metric, week_a_value, week_b_value, delta_value, metric_key in _weekly_compare_rows(comparison):
            hint = _comparison_hint(metric_key, delta_value)
            values = [metric, week_a_value, week_b_value, delta_value, hint]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = BORDER_ALL
                cell.alignment = ALIGN_CENTER if col > 1 else ALIGN_LEFT
                if col == 4:
                    if metric_key in {"high_risk_dogs", "moderate_risk_dogs", "underused_dogs"}:
                        cell.fill = FILL_GOOD if delta_value < 0 else FILL_WARN if delta_value > 0 else FILL_SUMMARY
                    else:
                        cell.fill = FILL_GOOD if delta_value > 0 else FILL_WARN if delta_value < 0 else FILL_SUMMARY
            row += 1

    ws.freeze_panes = "A4"
    _set_number_formats(ws)
    _autosize_columns(ws, min_width=14, max_width=30)


def _build_kennel_summary_sheet(wb: Workbook, db: Session, weekly, date_from: date, date_to: date) -> None:
    ws = wb.create_sheet("Kennel Summary")

    row = _write_title(
        ws,
        title="Kennel Summary",
        subtitle="Grouped by housing row and block. KM = kilometers, WORKED = unique worked dogs, STARTIT = safari count.",
        width_to_col=9 + len(weekly.items) * 3,
    )

    dogs = _fetch_operational_dogs(db)
    worklogs = _fetch_worklogs_in_range(db, date_from, date_to)
    dogs_by_id = {dog.id: dog for dog in dogs}
    week_order = [item.week_start for item in weekly.items]
    week_labels = {item.week_start: item.week_label for item in weekly.items}

    group_dogs: dict[tuple[str, int | None], set[int]] = defaultdict(set)
    for dog in dogs:
        group_dogs[(dog.kennel_row or "-", dog.kennel_block)].add(dog.id)

    by_group_week: dict[tuple[str, int | None], dict[date, dict[str, object]]] = defaultdict(
        lambda: defaultdict(lambda: {"km": 0.0, "kerrat": 0.0, "worked_dogs": set()})
    )

    for log in worklogs:
        dog = dogs_by_id.get(log.dog_id)
        if dog is None:
            continue
        week_start = _start_of_week(log.work_date)
        if week_start not in week_labels:
            continue
        group_key = (dog.kennel_row or "-", dog.kennel_block)
        bucket = by_group_week[group_key][week_start]
        bucket["km"] = float(bucket["km"]) + _to_float(log.km)
        bucket["kerrat"] = float(bucket["kerrat"]) + int(log.programs_3km or 0) + int(log.programs_10km or 0)
        if log.worked:
            cast_set: set[int] = bucket["worked_dogs"]  # type: ignore[assignment]
            cast_set.add(log.dog_id)

    base_headers = [
        "Housing row",
        "Block",
        "Dogs",
        "Total km",
        "Worked dogs",
        "Startit",
    ]

    col = 1
    for header in base_headers:
        top = ws.cell(row=row, column=col, value=header)
        bottom = ws.cell(row=row + 1, column=col, value="")
        for cell in (top, bottom):
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_ALL
        col += 1

    for index, week_start in enumerate(week_order):
        fill = _week_fill(index)
        start_col = col
        end_col = col + 2
        _merge_with_style(ws, row, start_col, row, end_col, week_labels[week_start], fill, FONT_HEADER, ALIGN_CENTER, BORDER_DARK)

        for offset, sub in enumerate(["KM", "WORKED", "STARTIT"]):
            c = ws.cell(row=row + 1, column=start_col + offset, value=sub)
            c.fill = fill
            c.font = FONT_HEADER
            c.alignment = ALIGN_CENTER
            c.border = BORDER_DARK
        col += 3

    row += 2
    ordered_groups = sorted(group_dogs.keys(), key=lambda x: ((x[0] or ""), x[1] if x[1] is not None else 999))

    for kennel_row, kennel_block in ordered_groups:
        total_km = 0.0
        total_kerrat = 0.0
        total_worked_dogs: set[int] = set()

        for week_start in week_order:
            stats = by_group_week[(kennel_row, kennel_block)][week_start]
            total_km += float(stats["km"])
            total_kerrat += float(stats["kerrat"])
            total_worked_dogs.update(stats["worked_dogs"])  # type: ignore[arg-type]

        fixed_values = [
            kennel_row,
            kennel_block,
            len(group_dogs[(kennel_row, kennel_block)]),
            round(total_km, 2),
            len(total_worked_dogs),
            int(total_kerrat),
        ]

        col = 1
        for value in fixed_values:
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = BORDER_ALL
            cell.alignment = ALIGN_CENTER
            cell.fill = FILL_SUMMARY
            col += 1

        for index, week_start in enumerate(week_order):
            fill = _week_fill(index)
            stats = by_group_week[(kennel_row, kennel_block)][week_start]
            week_values = [
                round(float(stats["km"]), 2),
                len(stats["worked_dogs"]),  # type: ignore[arg-type]
                int(float(stats["kerrat"])),
            ]
            for value in week_values:
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = BORDER_DARK
                cell.alignment = ALIGN_CENTER
                cell.fill = fill
                col += 1
        row += 1

    ws.freeze_panes = "A4"
    _set_number_formats(ws)
    _autosize_columns(ws, min_width=10, max_width=20)


def _build_dog_weekly_summary_sheet(wb: Workbook, db: Session, weekly, date_from: date, date_to: date) -> None:
    ws = wb.create_sheet("Dog Weekly Summary")

    week_order = [item.week_start for item in weekly.items]
    week_labels = {item.week_start: item.week_label for item in weekly.items}
    width_to_col = 1 + len(week_order) * 4 + 5

    row = _write_title(
        ws,
        title="Dog Weekly Summary",
        subtitle="Dogs grouped by housing blocks. Weekly columns: KM, STARTIT, TP, VP. Totals at right: KM, STARTIT, TP, VP, TP/VP.",
        width_to_col=width_to_col,
    )

    dogs = _fetch_operational_dogs(db)
    worklogs = _fetch_worklogs_in_range(db, date_from, date_to)

    by_dog_week: dict[int, dict[date, dict[str, float]]] = defaultdict(
        lambda: defaultdict(lambda: {"km": 0.0, "kerrat": 0.0, "tp": 0.0, "vp": 0.0})
    )
    dogs_worked_by_group_week: dict[tuple[str, int | None], dict[date, set[int]]] = defaultdict(
        lambda: defaultdict(set)
    )
    totals_by_group_week: dict[tuple[str, int | None], dict[date, dict[str, float]]] = defaultdict(
        lambda: defaultdict(lambda: {"km": 0.0, "kerrat": 0.0, "tp": 0.0, "vp": 0.0})
    )

    dogs_by_id = {dog.id: dog for dog in dogs}

    for log in worklogs:
        dog = dogs_by_id.get(log.dog_id)
        if dog is None:
            continue
        week_start = _start_of_week(log.work_date)
        if week_start not in week_labels:
            continue

        dog_bucket = by_dog_week[dog.id][week_start]
        dog_bucket["km"] += _to_float(log.km)
        dog_bucket["kerrat"] += int(log.programs_3km or 0) + int(log.programs_10km or 0)
        if log.worked:
            dog_bucket["tp"] += 1
        else:
            dog_bucket["vp"] += 1

        group_key = (dog.kennel_row or "-", dog.kennel_block)
        group_bucket = totals_by_group_week[group_key][week_start]
        group_bucket["km"] += _to_float(log.km)
        group_bucket["kerrat"] += int(log.programs_3km or 0) + int(log.programs_10km or 0)
        if log.worked:
            group_bucket["tp"] += 1
            dogs_worked_by_group_week[group_key][week_start].add(log.dog_id)
        else:
            group_bucket["vp"] += 1

    grouped_dogs: dict[tuple[str, int | None], list[Dog]] = defaultdict(list)
    for dog in dogs:
        grouped_dogs[(dog.kennel_row or "-", dog.kennel_block)].append(dog)

    ordered_groups = sorted(grouped_dogs.keys(), key=lambda x: ((x[0] or ""), x[1] if x[1] is not None else 999))

    dog_top = ws.cell(row=row, column=1, value="DOG")
    dog_bottom = ws.cell(row=row + 1, column=1, value="")
    for cell in (dog_top, dog_bottom):
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_DARK

    current_col = 2
    for index, week_start in enumerate(week_order):
        fill = _week_fill(index)
        _merge_with_style(ws, row, current_col, row, current_col + 3, week_labels[week_start], fill, FONT_HEADER, ALIGN_CENTER, BORDER_DARK)
        for offset, sub in enumerate(["KM", "STARTIT", "TP", "VP"]):
            c = ws.cell(row=row + 1, column=current_col + offset, value=sub)
            c.fill = fill
            c.font = FONT_HEADER
            c.alignment = ALIGN_CENTER
            c.border = BORDER_DARK
        current_col += 4

    _merge_with_style(ws, row, current_col, row, current_col + 4, "TOTAL", FILL_TOTAL, FONT_HEADER, ALIGN_CENTER, BORDER_DARK)
    for offset, sub in enumerate(["KM", "STARTIT", "TP", "VP", "TP/VP"]):
        c = ws.cell(row=row + 1, column=current_col + offset, value=sub)
        c.fill = FILL_TOTAL
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_DARK

    row += 2

    for kennel_row, kennel_block in ordered_groups:
        group_dogs = grouped_dogs[(kennel_row, kennel_block)]

        _merge_with_style(
            ws,
            row,
            1,
            row,
            width_to_col,
            f"Row {kennel_row} / Block {kennel_block}",
            FILL_SECTION,
            FONT_SECTION,
            ALIGN_LEFT,
            BORDER_DARK,
        )
        row += 1

        for dog in group_dogs:
            name_cell = ws.cell(row=row, column=1, value=dog.name)
            name_cell.fill = FILL_NAME
            name_cell.font = FONT_BOLD
            name_cell.alignment = ALIGN_LEFT
            name_cell.border = BORDER_DARK

            current_col = 2
            total_km = 0.0
            total_kerrat = 0.0
            total_tp = 0.0
            total_vp = 0.0

            for index, week_start in enumerate(week_order):
                fill = _week_fill(index)
                stats = by_dog_week[dog.id][week_start]
                total_km += stats["km"]
                total_kerrat += stats["kerrat"]
                total_tp += stats["tp"]
                total_vp += stats["vp"]

                week_values = [
                    round(stats["km"], 2),
                    int(stats["kerrat"]),
                    int(stats["tp"]),
                    int(stats["vp"]),
                ]
                for value in week_values:
                    cell = ws.cell(row=row, column=current_col, value=value)
                    cell.fill = fill
                    cell.border = BORDER_DARK
                    cell.alignment = ALIGN_CENTER
                    current_col += 1

            total_values = [
                round(total_km, 2),
                int(total_kerrat),
                int(total_tp),
                int(total_vp),
                _safe_ratio(total_tp, total_vp),
            ]
            for value in total_values:
                cell = ws.cell(row=row, column=current_col, value=value)
                cell.fill = FILL_TOTAL
                cell.border = BORDER_DARK
                cell.alignment = ALIGN_CENTER
                current_col += 1

            row += 1

        total_label = ws.cell(row=row, column=1, value="Group totals")
        total_label.fill = FILL_SUMMARY
        total_label.font = FONT_BOLD
        total_label.alignment = ALIGN_LEFT
        total_label.border = BORDER_DARK

        current_col = 2
        group_total_km = 0.0
        group_total_kerrat = 0.0
        group_total_tp = 0.0
        group_total_vp = 0.0

        for index, week_start in enumerate(week_order):
            fill = _week_fill(index)
            stats = totals_by_group_week[(kennel_row, kennel_block)][week_start]
            group_total_km += stats["km"]
            group_total_kerrat += stats["kerrat"]
            group_total_tp += stats["tp"]
            group_total_vp += stats["vp"]

            for value in [
                round(stats["km"], 2),
                int(stats["kerrat"]),
                int(stats["tp"]),
                int(stats["vp"]),
            ]:
                cell = ws.cell(row=row, column=current_col, value=value)
                cell.fill = fill
                cell.border = BORDER_DARK
                cell.alignment = ALIGN_CENTER
                current_col += 1

        for value in [
            round(group_total_km, 2),
            int(group_total_kerrat),
            int(group_total_tp),
            int(group_total_vp),
            _safe_ratio(group_total_tp, group_total_vp),
        ]:
            cell = ws.cell(row=row, column=current_col, value=value)
            cell.fill = FILL_TOTAL
            cell.border = BORDER_DARK
            cell.alignment = ALIGN_CENTER
            current_col += 1

        row += 1

        worked_label = ws.cell(row=row, column=1, value="Dogs worked")
        worked_label.fill = FILL_SUMMARY
        worked_label.font = FONT_BOLD
        worked_label.alignment = ALIGN_LEFT
        worked_label.border = BORDER_DARK

        current_col = 2
        for index, week_start in enumerate(week_order):
            fill = _week_fill(index)
            worked_count = len(dogs_worked_by_group_week[(kennel_row, kennel_block)][week_start])
            for value in ["", "", worked_count, ""]:
                cell = ws.cell(row=row, column=current_col, value=value)
                cell.fill = fill
                cell.border = BORDER_DARK
                cell.alignment = ALIGN_CENTER
                current_col += 1

        for _ in range(5):
            cell = ws.cell(row=row, column=current_col, value=None)
            cell.fill = FILL_TOTAL
            cell.border = BORDER_DARK
            current_col += 1

        row += 2

    ws.freeze_panes = "B4"
    _set_number_formats(ws)
    _autosize_columns(ws, min_width=10, max_width=18)


def build_raw_run_logs_csv(db: Session, date_from: date, date_to: date):
    dogs = {dog.id: dog for dog in _fetch_operational_dogs(db)}
    worklogs = _fetch_worklogs_in_range(db, date_from, date_to)

    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow([
        "work_date",
        "week_label",
        "dog_name",
        "kennel_row",
        "kennel_block",
        "home_slot",
        "primary_role",
        "worked",
        "km",
        "programs_3km",
        "programs_10km",
        "kerrat_total",
        "status",
        "notes",
    ])

    for log in worklogs:
        dog = dogs.get(log.dog_id)
        writer.writerow([
            log.work_date.isoformat(),
            log.week_label,
            dog.name if dog else f"dog_id={log.dog_id}",
            log.kennel_row or (dog.kennel_row if dog else None),
            dog.kennel_block if dog else None,
            log.home_slot or (dog.home_slot if dog else None),
            log.main_role or (dog.primary_role if dog else None),
            "yes" if log.worked else "no",
            round(_to_float(log.km), 2),
            int(log.programs_3km or 0),
            int(log.programs_10km or 0),
            int(log.programs_3km or 0) + int(log.programs_10km or 0),
            log.status,
            log.notes,
        ])

    content = BytesIO(buffer.getvalue().encode("utf-8-sig"))
    content.seek(0)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    filename = f"husky-run-logs-{date_from.isoformat()}-to-{date_to.isoformat()}-{timestamp}.csv"
    return content, filename


def build_analytics_summary_pdf(
    db: Session,
    date_from: date,
    date_to: date,
    week_a_start: date | None = None,
    week_b_start: date | None = None,
):
    weekly = get_weekly_analytics(db=db, date_from=date_from, date_to=date_to)
    summary = get_analytics_summary(db=db, date_from=date_from, date_to=date_to)
    comparison = None
    if week_a_start is not None and week_b_start is not None:
        comparison = get_weekly_compare(db=db, week_a_start=week_a_start, week_b_start=week_b_start)

    labels = [item.week_start.strftime("%d.%m") for item in weekly.items]
    total_km = [item.total_km for item in weekly.items]
    worked_dogs = [item.worked_dogs for item in weekly.items]
    high_risk = [item.high_risk_dogs for item in weekly.items]
    moderate_risk = [item.moderate_risk_dogs for item in weekly.items]
    underused = [item.underused_dogs for item in weekly.items]
    avg_km = [item.avg_km_per_worked_dog for item in weekly.items]

    stream = BytesIO()
    with PdfPages(stream) as pdf:
        fig = plt.figure(figsize=(11.69, 8.27))
        fig.patch.set_facecolor("white")

        ax = fig.add_axes([0, 0, 1, 1])
        ax.axis("off")

        ax.text(0.05, 0.94, "Husky Tracking AI", fontsize=24, fontweight="bold")
        ax.text(0.05, 0.90, "Analytics Summary Report", fontsize=18)
        ax.text(
            0.05,
            0.865,
            f"Period: {date_from.isoformat()} -> {date_to.isoformat()} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            fontsize=10,
            color="#475569",
        )

        cards = [
            ("Total km", f"{summary.total_km:.0f}"),
            ("Worked dogs", f"{summary.unique_worked_dogs}"),
            ("Avg km / worked dog", f"{summary.avg_km_per_worked_dog:.2f}"),
            (
                "Latest snapshot",
                (
                    f"H {summary.latest_week_snapshot.high_risk_dogs} / "
                    f"M {summary.latest_week_snapshot.moderate_risk_dogs} / "
                    f"U {summary.latest_week_snapshot.underused_dogs}"
                ) if summary.latest_week_snapshot else "N/A",
            ),
        ]

        x_positions = [0.05, 0.285, 0.52, 0.755]
        for (title, value), x in zip(cards, x_positions):
            ax.add_patch(Rectangle((x, 0.73), 0.19, 0.10, facecolor="#F8FAFC", edgecolor="#CBD5E1"))
            ax.text(x + 0.015, 0.80, title, fontsize=11, color="#64748B")
            ax.text(x + 0.015, 0.755, value, fontsize=20, fontweight="bold", color="#0F172A")

        ax_bar1 = fig.add_axes([0.07, 0.42, 0.38, 0.22])
        ax_bar1.bar(labels, total_km)
        ax_bar1.set_title("Total km by week", fontsize=12, fontweight="bold")
        ax_bar1.tick_params(axis="x", rotation=0, labelsize=9)
        ax_bar1.tick_params(axis="y", labelsize=9)

        ax_bar2 = fig.add_axes([0.55, 0.42, 0.38, 0.22])
        ax_bar2.bar(labels, worked_dogs)
        ax_bar2.set_title("Worked dogs by week", fontsize=12, fontweight="bold")
        ax_bar2.tick_params(axis="x", rotation=0, labelsize=9)
        ax_bar2.tick_params(axis="y", labelsize=9)

        ax_line1 = fig.add_axes([0.07, 0.10, 0.86, 0.20])
        ax_line1.plot(labels, high_risk, marker="o", label="High risk")
        ax_line1.plot(labels, moderate_risk, marker="o", label="Moderate risk")
        ax_line1.plot(labels, underused, marker="o", label="Underused")
        ax_line1.set_title("Risk and underuse trend", fontsize=12, fontweight="bold")
        ax_line1.legend(fontsize=9)
        ax_line1.tick_params(axis="x", labelsize=9)
        ax_line1.tick_params(axis="y", labelsize=9)

        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        fig2 = plt.figure(figsize=(11.69, 8.27))
        fig2.patch.set_facecolor("white")
        ax2 = fig2.add_axes([0, 0, 1, 1])
        ax2.axis("off")

        ax2.text(0.05, 0.94, "Weekly efficiency and comparison", fontsize=20, fontweight="bold")

        ax_avg = fig2.add_axes([0.08, 0.56, 0.84, 0.24])
        ax_avg.plot(labels, avg_km, marker="o")
        ax_avg.set_title("Average km per worked dog", fontsize=12, fontweight="bold")
        ax_avg.tick_params(axis="x", labelsize=9)
        ax_avg.tick_params(axis="y", labelsize=9)

        if comparison is not None:
            ax2.text(0.05, 0.46, f"Week comparison: {comparison.week_a.week_label} vs {comparison.week_b.week_label}", fontsize=14, fontweight="bold")
            compare_rows = []
            for metric, week_a_value, week_b_value, delta_value, metric_key in _weekly_compare_rows(comparison):
                compare_rows.append([metric, week_a_value, week_b_value, delta_value, _comparison_hint(metric_key, delta_value)])

            table = ax2.table(
                cellText=compare_rows,
                colLabels=["Metric", "Week A", "Week B", "Change", "Interpretation"],
                cellLoc="center",
                colLoc="center",
                bbox=[0.05, 0.12, 0.90, 0.26],
            )
            table.auto_set_font_size(False)
            table.set_fontsize(10)
            for (r, c), cell in table.get_celld().items():
                if r == 0:
                    cell.set_text_props(weight="bold")
                    cell.set_facecolor("#CBD5E1")
                else:
                    cell.set_facecolor("#F8FAFC")
        else:
            ax2.text(0.05, 0.40, "No week comparison selected.", fontsize=12)

        pdf.savefig(fig2, bbox_inches="tight")
        plt.close(fig2)

    stream.seek(0)
    return stream


def build_analytics_workbook(
    db: Session,
    date_from: date,
    date_to: date,
    week_a_start: date | None = None,
    week_b_start: date | None = None,
):
    weekly = get_weekly_analytics(db=db, date_from=date_from, date_to=date_to)
    summary = get_analytics_summary(db=db, date_from=date_from, date_to=date_to)

    comparison = None
    if week_a_start is not None and week_b_start is not None:
        comparison = get_weekly_compare(db=db, week_a_start=week_a_start, week_b_start=week_b_start)

    wb = Workbook()
    _build_overview_sheet(wb=wb, date_from=date_from, date_to=date_to, weekly=weekly, summary=summary, comparison=comparison)
    _build_weekly_comparison_sheet(wb=wb, comparison=comparison)
    _build_kennel_summary_sheet(wb=wb, db=db, weekly=weekly, date_from=weekly.date_from, date_to=weekly.date_to)
    _build_dog_weekly_summary_sheet(wb=wb, db=db, weekly=weekly, date_from=weekly.date_from, date_to=weekly.date_to)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream